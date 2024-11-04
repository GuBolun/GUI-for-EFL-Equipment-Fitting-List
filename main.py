###（导出数据）子进程相关配置和函数
from openpyxl import styles,load_workbook

def items_order(order_dict,to_sort_list):
    """将乱序items按照模板配件顺序排列"""
    in_order = []
    not_in_order = []
    for item in to_sort_list:
        if (temp:=item[0].split('-'))[0] in order_dict:
            in_order.append(tuple(temp)+(item[1],item[2],item[1]*item[2]))
        else:
            not_in_order.append(tuple(temp)+(item[1],item[2],item[1]*item[2]))
    in_order.sort(key=lambda x: order_dict[x[0]])
    sorted_list = in_order + not_in_order
    return sorted_list

def output_p(output_event,results,template_path,order_H,order_J,order_Z):
    """将所有数据根据在线模板'配件明细汇总_.xlsx'/本地文件'配件明细汇总导出.xlsx'导出到本地文件'配件明细汇总导出.xlsx'"""
    output_wb=load_workbook(template_path)
    ws_all=output_wb['汇总']
    for res in results:
        num,vtype,addremark_dict=0,'',{}#序号
        for i in range(2,ws_all.max_row+1):#查找是否已经有记录。i是从1开始的行号，为了下一行if判断中减少计算
            if ws_all['B'+str(i)].value==res[0] and ws_all['C'+str(i)].value==res[1]:
                num=i-1#行号转序号
                ws_all['D'+str(i)]=res[2]#找到记录，更新成本
                ws_all['A'+str(i)]=f'=HYPERLINK("#{num}!A1","{num}")'
                ws_new=output_wb[str(num)]
                order={}
                for idx, val in enumerate(tuple([row[0] for row in ws_new.iter_rows(min_row=6, max_col=1, values_only=True) if row[0] is not None][:-1])):
                    order[val] = 2 * idx
                    order[val+'成品'] = 2 * idx + 1 ##即使出现成品成品，顺序也是对的！
                if ws_new['A2'].value == '截止阀':
                    vtype='J'
                elif ws_new['A2'].value == '闸阀':
                    vtype='Z'
                elif ws_new['A2'].value == '止回阀':
                    vtype='H'
                else:
                    vtype='J'
                for i in range(6,ws_new.max_row+1):##遍历保存手动添加的备注
                    if (value_:=ws_new.cell(i,9).value) is not None:
                        addremark_dict[ws_new.cell(i,1).value]=value_
                break
        if num==0:#未找到记录，则在末尾新增
            num=ws_all.max_row#序号
            ws_all.append((num,res[0],res[1],res[2]))
            ws_all['A'+str(num+1)]=f'=HYPERLINK("#{num}!A1","{num}")'
            vtype_=res[0].split('-')[0]
            if 'J' in vtype_:
                vtype='J'
                ws_new=output_wb.copy_worksheet(output_wb['截止阀模板'])
                order=order_J
            elif 'Z' in vtype_:
                vtype='Z'
                ws_new=output_wb.copy_worksheet(output_wb['闸阀模板'])
                order=order_Z
            elif 'H' in vtype_:
                vtype='H'
                ws_new=output_wb.copy_worksheet(output_wb['止回阀模板'])
                order=order_H
            else:##按相对较全的截止阀模板导出
                vtype='J'
                ws_new=output_wb.copy_worksheet(output_wb['截止阀模板'])
                order=order_J
        ws_new.title=str(num)
        ws_new['A1']=f'=HYPERLINK("#汇总!A{num+1}","{num}")'
        ws_new['B1']=res[0]#型号名
        ws_new['B2']=res[2]#成本
        ws_new['D2']=res[5]#炉号
        ws_new['E2']=res[4]#成品重量
        ws_new['F2']=res[3]#日期
        ws_new['B4']=res[1]#毛坯货源
        ws_new['C4']=res[6]#毛坯重量
        ws_new['D4']=res[7]#毛坯单价
        if res[6]==0:
            ws_new['E4']=res[7]#毛坯小计
        else:
            ws_new['E4']=res[6]*res[7]#毛坯小计
        ws_new['F4']=res[11]#电焊
        ws_new['G4']=res[8]#加工费
        newremark_=res[12]#备注
        ind=newremark_.find('；')
        newremark_=newremark_[ind+1:]
        if vtype != 'H':
            ind=newremark_.find('阀杆Tr')#res[12]#备注
            if ind == -1:
                ws_new['G2']=f'{res[9]}*{res[10]}'#阀杆
            else:
                ind2=newremark_[ind:].find('。')
                temp_=newremark_[ind+2:ind+ind2]
                newremark_=newremark_[:ind]+newremark_[ind+ind2+1:]
                ws_new['G2']=f'{res[9]}*{res[10]} {temp_}'#阀杆
        
        remark_dict={key:'' for key in order}
        remain_remark=[]
        for ir in newremark_.split('。'):
            flag=0
            for k in remark_dict:
                if ir.startswith(k):
                    remark_dict[k]=ir
                    flag=1
                    break
            if not flag:
                remain_remark.append(ir)
        remain_remark='。'.join(remain_remark)

        order_items=items_order(order,res[13])
        default_font=styles.Font(size=14)
        cost_, i = 0, -1
        for i,item in enumerate(order_items):
            cost_ += item[5]
            for j in range(6):
                ws_new.cell(i+6,j+1,item[j]).font=default_font
            ws_new.cell(i+6,7,remark_dict.pop(item[0],'')).font=default_font
            ws_new.cell(i+6,9,addremark_dict.pop(item[0],'')).font=default_font
        ws_new.cell(i+8,9,addremark_dict.pop('合计','')).font=default_font
        pre=str(v_)+'。' if (v_:=ws_new.cell(i+7,9).value) is not None else ''
        ws_new.cell(i+7,9,(pre+'。'.join(map(str, addremark_dict.values()))).replace('。。','。')).font=default_font#无法匹配的备注统一放在后面
        for r in range(i+7,i+10):
            for j in range(1,7):
                ws_new.cell(r,j,'').font=default_font
        ws_new.cell(i+9,1,'合计').font=default_font
        ws_new.cell(i+9,6,res[2]).font=default_font
        ws_new.cell(i+9,7,remain_remark).font=default_font
        for j in range(i,len(order)//2-1):###删去总计下面多余的模板配件
            ws_new.cell(j+10,1,'').font=default_font
        ws_new['H4']=res[2]-cost_#小计
        rodcost_=ws_new['H4'].value-ws_new['E4'].value-ws_new['F4'].value-ws_new['G4'].value#阀杆价格
        if (rr:=round(rodcost_,3)) != 0:
            ws_new['H1']=rr#阀杆价格
        else:
            ws_new['H1']=''
    try:
        output_wb.save(f"配件明细汇总导出.xlsx")
        output_event.put(1)
    except PermissionError:
        output_event.put(0)

if __name__ == '__main__': ##避免multiprocessing加载不必要的全局变量
    from os.path import isfile
    import sqlite3
    from tkinter import Tk,ttk,StringVar,messagebox,Toplevel,Listbox,Label,Entry,Frame,Button,Menu,Text,Scrollbar,END,BOTH,WORD,NO,YES,INSERT
    from tkinter.font import Font,nametofont
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    from pandas import read_excel
    import datetime
    import re
    from multiprocessing import Process,Queue,freeze_support
    from functools import cmp_to_key
    freeze_support()###使打包后的exe文件在拉起新的进程时不会重新执行if __name__ == '__main__'下的内容！！！

    database_path="main.db"
    config_path=r".\config.xlsx"
    config_path_="."
    log_path="log.txt"

    template_path=r"配件明细汇总_.xlsx"
    items_path=r"配件价格.xlsx"
    Today = str(datetime.date.today())

    ####数据库基本操作
    class TTVALVE:
        """创建数据库类，直接管理数据库。不过仍未实现数据库与界面操作的完全独立。"""
        def __init__(self):
            self.conn = sqlite3.connect(database_path)
            self.cur = self.conn.cursor()
            self.execute(
                "CREATE TABLE IF NOT EXISTS valves (name TEXT PRIMARY KEY, date DATE, weight REAL, hnumber TEXT, rodd TEXT, rodl TEXT , diff REAL, remark TEXT)")
            self.execute(
                "CREATE TABLE IF NOT EXISTS items (iname TEXT PRIMARY KEY, icost REAL)")
            self.execute(
                "CREATE TABLE IF NOT EXISTS relationships (vname TEXT, iname TEXT, count REAL, PRIMARY KEY (vname,iname))")
            self.execute(
                "CREATE TABLE IF NOT EXISTS re_supplies (name TEXT, supply TEXT, cost REAL, bweight REAL, bwcost REAL, process REAL, PRIMARY KEY (name,supply))")
            self.conn.commit()

        def __del__(self):
            self.conn.close()

        def execute(self, sql, parameters=(), safe=True):
            if safe:
                try:
                    self.cur.execute(sql, parameters)
                except Exception as e:
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    with open(log_path,"a") as logfile:
                        logfile.write(f"{timestamp}: 错误:{e}\t尝试执行'{sql}'\t使用的参数{parameters}\n")
                    messagebox.showerror("错误",f"意料之外的错误{e}，建议重启软件后重试。")
            else:
                self.cur.execute(sql, parameters)

        def executemany(self, sql, seq_of_parameters=(), safe=True):
            if safe:
                try:
                    self.cur.executemany(sql, seq_of_parameters)
                except Exception as e:
                    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    with open(log_path,"a") as logfile:
                        logfile.write(f"{timestamp}: 错误:{e}\t尝试执行'{sql}'\t使用的参数{seq_of_parameters}\n")
                    messagebox.showerror("错误",f"意料之外的错误{e}")
            else:
                self.cur.executemany(sql, seq_of_parameters)

        def insert_valve(self, name, supply=None, cost=0.0, date=Today, weight=0.0, hnumber='', bweight=0.0, bwcost=0.0, process=0.0, rodd='', rodl='', diff=0.0, remark=''):
            self.execute("INSERT INTO valves VALUES (?,?,?,?,?,?,?,?)", (name, date, weight, hnumber, rodd, rodl, diff, remark))
            if supply is None:
                self.execute("INSERT INTO re_supplies VALUES (?,?,?,?,?,?)", (name, '未指定', 0, 0, 0, 0))
            else:
                self.execute("INSERT INTO re_supplies VALUES (?,?,?,?,?,?)", (name, supply, cost, bweight, bwcost, process))
            self.conn.commit()
            # self.cal_cost_(name,supply)   #插入伴随着显示，显示中已包含了成本的更新

        def search_valve(self, name, supply=''):
            self.execute("SELECT * FROM valves WHERE name=?", (name,))
            res1 = self.cur.fetchone()
            if supply == '':
                self.execute("SELECT * FROM re_supplies WHERE name=? ORDER BY cost", (name,))
                res2 = self.cur.fetchone()
            else:
                self.execute("SELECT * FROM re_supplies WHERE name=? and supply=?", (name,supply))
                res2 = self.cur.fetchone()
            if res1 is None or res2 is None:
                return None
            return (res1[0],res2[1],res2[2],res1[1],res1[2],res1[3],res2[3],res2[4],res2[5],res1[4],res1[5],res1[6],res1[7])
        
        def search_items(self, name):
            """return [(iname_type,iname_spe,iname_tex,count,icost),]"""
            self.execute("SELECT items.iname,relationships.count,items.icost FROM relationships JOIN items on relationships.iname=items.iname where relationships.vname=?", (name,))
            res = self.cur.fetchall()
            vtype_=name.split('-')[0]
            if 'J' in vtype_:
                order = order_J
            elif 'Z' in vtype_:
                order = order_Z
            elif 'H' in vtype_:
                order = order_H
            else:
                order = order_J
            return items_order(order,res)

        def update_valve(self, name, supply='未指定', cost=0.0, date=Today , weight=0.0, hnumber='', bweight=0.0, bwcost=0.0, process=0.0, rodd='', rodl='', diff=0.0, remark=''):#默认值仅展示，实际不起作用，因为即使传入None也无法赋默认值
            self.execute("UPDATE valves SET date=?, weight=?, hnumber=?, rodd=?, rodl=?, diff=?, remark=? WHERE name=?", (date, weight, hnumber, rodd, rodl, diff, remark, name))
            self.conn.commit()
            self.insert_supply(name, supply, cost, bweight, bwcost, process)
            self.cal_cost_(name,supply,True)

        def insert_item(self, iname, icost):#含update
            self.execute("INSERT INTO items VALUES (?,?) ON CONFLICT(iname) DO UPDATE SET icost=EXCLUDED.icost", (iname,icost))
            itype,ispe,itex=iname.split('-')
            if itype == '石墨圈':
                self.execute("UPDATE items SET icost=? WHERE iname LIKE ?", (icost,itype+'-%-'+itex))
            elif itype == '铜螺母':
                self.execute("UPDATE items SET icost=? WHERE iname LIKE ?", (icost,itype+'-%'))
            elif itype[-2:] == '螺帽':
                self.execute("UPDATE items SET icost=? WHERE iname LIKE ?", (icost,f'%螺帽-{ispe}-{itex}'))
            elif itype == '压子':
                spe=ispe.split('*')[0]
                self.execute("UPDATE items SET icost=? WHERE iname LIKE ?",(icost,f'{itype}-{spe}%-{itex}'))
            elif itex in config_texlist and itype in config_typeslist:# and itype in config_typeslist加的话相当于限制散件只能是规定的
                param=[ (icost,type+'-%-'+itex) for type in config_typeslist]
                self.executemany("UPDATE items SET icost=? WHERE iname LIKE ?", param)
            self.conn.commit()

        def insert_items(self, items, adds):
            self.executemany("INSERT INTO items VALUES (?,?) ON CONFLICT(iname) DO UPDATE SET icost=EXCLUDED.icost", items)
            self.executemany("UPDATE items SET icost=? WHERE iname LIKE ?", adds)
            self.conn.commit()

        def search_item(self, iname):
            self.execute("SELECT * FROM items WHERE iname=?", (iname,))
            res = self.cur.fetchone()
            return res

        def insert_relationship(self, name, iname, count=0):##含update
            if count == 0 or count == '0':
                messagebox.showinfo("提示","配件数量不能为空！")
                return
            if self.search_valve(name) is None:
                messagebox.showinfo("提示","型号不存在添加失败！")
                return
            self.execute("INSERT INTO relationships VALUES (?,?,?) ON CONFLICT(vname,iname) DO UPDATE SET count=EXCLUDED.count", (name,iname,count))
            self.conn.commit()
            self.conn.commit()
            self.cal_cost_(name)

        def insert_relationships_nocheck(self, relationships):
            self.executemany("INSERT INTO relationships VALUES (?,?,?)", relationships)
            self.conn.commit()     

        def insert_item_relationship(self, name, iname, icost=0, count=0):
            self.insert_item(iname,icost)
            self.insert_relationship(name,iname,count)

        def search_relationship(self, name, iname):
            self.execute("SELECT * FROM relationships WHERE vname=? and iname=?", (name,iname))
            res = self.cur.fetchone()
            return res
        
        def insert_supply(self, name, supply, cost, bweight, bwcost, process):##含update
            self.execute("""INSERT INTO re_supplies VALUES (?,?,?,?,?,?) on CONFLICT(name,supply) 
                DO UPDATE SET cost=EXCLUDED.cost,bweight=EXCLUDED.bweight,
                bwcost=EXCLUDED.bwcost,process=EXCLUDED.process""",(name,supply,cost,bweight,bwcost,process))
            self.conn.commit()
            if supply != '未指定':
                #回写到excel中
                global config_bwcost
                new_config_bwcost = config_bwcost.copy()
                new_config_bwcost[supply] = float(bwcost)
                if config_bwcost != new_config_bwcost:
                    config_bwcost = new_config_bwcost
                    sort = list(config_bwcost.keys())
                    try:
                        ind = sort.index(supply)+2
                    except ValueError:
                        ind = len(sort)+2
                    wb = load_workbook(config_path)
                    wb['毛坯']['A'+str(ind)]=supply
                    wb['毛坯']['B'+str(ind)]=float(bwcost)
                    write_to_excel(wb)

        def reset_supplies(self, supplies, name):
            try:
                self.execute("DELETE FROM re_supplies WHERE name=?",(name,),safe=False)
                for supply in supplies:
                    self.insert_supply(*supply)
                if not self.search_supplies(name):
                    self.execute("INSERT INTO re_supplies VALUES (?,'未指定',0,0,0,0)",(name,),safe=False)
                self.execute("UPDATE valves SET date=? WHERE name=?",(Today, name),safe=False)
            except Exception as e:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                with open(log_path,"a") as logfile:
                    logfile.write(f"{timestamp}: reset_supplies错误:{e}\n")
                return False
            else:
                self.conn.commit()
                return True

        def search_supplies(self, name):
            self.execute("SELECT * FROM re_supplies WHERE name=? ORDER BY cost", (name,))
            res = self.cur.fetchall()
            return res

        def update_supplies(self):
            for supply,bwcost in config_bwcost.items():
                bwcost=float(bwcost)
                self.execute("UPDATE re_supplies SET bwcost=? where supply=?",(bwcost,supply))
                self.conn.commit()
                        
        def delete_relationship(self, name, iname, count):
            res=self.search_relationship(name,iname)
            if res is None:
                messagebox.showinfo("提示",f"'{name}'型号中不存在'{iname}'配件，请检查！")
                return False
            if res[2] != float(count):
                messagebox.showwarning("警告",f"输入的配件数量'{count}'不等于'{name}'型号中'{iname}'配件的数量'{res[2]}'，请检查！")
                return False
            try:
                self.execute("DELETE FROM relationships WHERE vname=? and iname=?", (name,iname),safe=False)
            except Exception as e:
                messagebox.showerror("错误",f"删除型号'{iname}'中的配件'{iname}'，数量{count}失败。")
                return False
            else:
                self.conn.commit()
                self.cal_cost_(name)
                return True

        def delete_valve(self, name):
            res = self.search_valve(name)
            if res is None:
                messagebox.showinfo("提示",f"'{name}'型号不存在，请检查！")
                return
            self.execute("DELETE FROM valves WHERE name=?", (name,))
            self.execute("DELETE FROM relationships WHERE vname=?", (name,))
            self.execute("DELETE FROM re_supplies WHERE name=?", (name,))
            self.conn.commit()

        def delete_item(self, iname):
            res = self.search_item(iname)
            if res is None:
                messagebox.showinfo("提示",f"配件'{iname}'不存在，请检查！")
                return
            self.execute("SELECT vname FROM relationships where iname=?", (iname,))
            results = self.cur.fetchall()
            if not results:
                self.execute("DELETE FROM items WHERE iname=?", (iname,))
                self.conn.commit()
                return
            namelist = [res[0] for res in results]
            ans = messagebox.askyesnocancel("删除配件确认",f"当前配件仍存在于{','.join(namelist)}。\n按是依然删除，按否修改当前配件名称。")
            if ans is None:
                return
            elif ans:
                self.execute("DELETE FROM items WHERE iname=?", (iname,))
                self.execute("DELETE FROM relationships WHERE iname=?", (iname,))
                self.conn.commit()
            else:
                window.wait_window(inputDialog(window,'修改配件名',f"将当前配件名称'{iname}'修改为：",set_callback_value))
                try:
                    newname = value_callback##使用全局变量做回调，谨慎使用
                except NameError:#防止用户直接关闭子窗口而报错
                    return
                if newname != '':
                    try:
                        self.execute("UPDATE items SET iname=? WHERE iname=? ", (newname,iname), safe=False)
                        self.execute("UPDATE relationships SET iname=? WHERE iname=? ", (newname,iname), safe=False)
                    except Exception as E:
                        if messagebox.askokcancel('警告','要修改的配件名已存在，是否合并'):
                            self.execute("SELECT vname, SUM(count) AS newcount FROM relationships where iname IN (?,?) GROUP BY vname", (newname,iname))
                            results2=self.cur.fetchall()
                            params=[(res[0], newname, res[1], res[1]) for res in results2]
                            self.execute("DELETE FROM items where iname=?",(iname,))
                            self.execute("DELETE FROM relationships where iname=?",(iname,))
                            self.executemany("INSERT INTO relationships VALUES (?,?,?) ON CONFLICT(vname,iname) DO UPDATE SET count=?", params)
                            self.conn.commit()
                    else:
                        self.conn.commit()

        def delete_supply(self, name, supply):
            res = self.search_valve(name,supply)
            if res is None:
                messagebox.showinfo("提示",f"'{name}'的货源'{supply}'不存在，无法删除，请检查。")
                return
            self.execute("DELETE FROM re_supplies WHERE name=? and supply=?", (name, supply))
            if not self.search_supplies(name):
                self.execute("INSERT INTO re_supplies VALUES (?,'未指定',0,0,0,0)",(name,))
            self.conn.commit()

        def fsearch_valvename(self, prename):
            if prename=='*':
                self.execute("SELECT name FROM valves")
            else:
                self.execute("SELECT name FROM valves WHERE name LIKE ?", (prename.replace('-','%-')+'%',))
            results = self.cur.fetchall()
            return [res[0] for res in results]

        def fsearch_itemname(self, prename):
            itype,ispe,itex=prename.split('-')[:3]
            selected_iname=f'{itype}%-{ispe}%-{itex}%'
            self.execute("SELECT iname FROM items WHERE iname LIKE ?", (selected_iname,))
            results = self.cur.fetchall()
            return [res[0] for res in results]

        def fsearch_item_valves(self, subiname):
            self.execute("SELECT * FROM items WHERE iname like ?", (f'%{subiname}%',))
            results=[]
            for res in self.cur.fetchall():
                self.execute("SELECT vname from relationships WHERE iname=?",(res[0],))
                vname_list=[row[0] for row in self.cur.fetchall()]
                results.append((res[0],res[1],','.join(vname_list)))
            return results

        def delete_trash(self):
            self.execute("BEGIN")
            self.execute("""
                SELECT r1.name, r1.supply FROM re_supplies r1
                WHERE r1.cost = 0 AND r1.bweight = 0 AND r1.bwcost = 0 AND r1.process = 0
                AND EXISTS (
                SELECT 1  FROM re_supplies r2
                WHERE r2.name = r1.name AND r2.supply <> r1.supply)
            """)###找到无内容货源
            supply_trash = self.cur.fetchall()
            self.executemany("DELETE FROM re_supplies WHERE name=? and supply=?",(supply_trash))
            self.execute("""SELECT v.name FROM valves v JOIN re_supplies rs ON rs.name = v.name
                WHERE rs.cost = 0 AND v.remark = ''  
                AND NOT EXISTS (SELECT 1 FROM relationships r WHERE r.vname = rs.name);""")###找到无内容型号
            valves_trash = self.cur.fetchall()
            self.executemany("DELETE FROM valves WHERE name=?",(valves_trash))
            self.executemany("DELETE FROM relationships WHERE vname=?",(valves_trash))
            self.executemany("DELETE FROM re_supplies WHERE name=?", (valves_trash))
            self.execute("""SELECT iname FROM items WHERE NOT EXISTS 
                        (SELECT 1 FROM relationships WHERE relationships.iname = items.iname)""")###找到无主配件
            results2 = self.cur.fetchall()
            items_noowner = [res[0] for res in results2]
            items_reserved = [f'--{x}' for x in config_texlist]+['石墨圈--','铜螺母--']
            items_trash = [(x,) for x in items_noowner if x not in items_reserved]
            self.executemany("DELETE FROM items WHERE iname=?",(items_trash))
            items_trash_=[x[0] for x in items_trash]
            if len(items_trash_) < 10:
                items_slice=','.join(items_trash_)
            else:
                items_slice=','.join(items_trash_[:5]) + ",...," + ','.join(items_trash_[-5:])
            if len(valves_trash) + len(items_trash) + len(supply_trash) == 0:
                messagebox.showinfo("清理","清理完成！")
                self.conn.commit()
            elif messagebox.askokcancel("清理", f"确认清理以下货源，阀门和配件?\n货源：{','.join([f'{t[0]}|{t[1]}' for t in supply_trash])}\n阀门：{','.join(x[0] for x in valves_trash)}\n配件：{items_slice}"):
                self.conn.commit()
            else:
                self.conn.rollback()
            self.execute("VACUUM")#重新组织数据库文件，减少碎片化，并优化数据库性能。

        def cal_cost_(self, name, supply=None, update_date=False):
            if supply == None:
                supply = E_supply.get()
            res = self.search_valve(name, supply)
            if res is None:
                return
            showdic['name'],showdic['supply'],showdic['process'],showdic['rodd'],showdic['diff'] = res[0],res[1],res[8],res[9],res[11]
            try:
                showdic['rodl'] = eval(res[10])
            except SyntaxError:
                showdic['rodl'] = ''
            former_cost,cost = res[2],0
            for item in self.search_items(name):
                cost += item[3] * item[4]
            showdic['items'] = cost
            showdic['rod'] = cal_rod(res[0],res[9],res[10],False)
            showdic['bw'] = res[7] if res[6]==0 else res[6] * res[7]
            showdic['cost'] = round(showdic['bw'] + showdic['process'] + showdic['rod'] + showdic['diff'] + showdic['items'],3)
            cost = showdic['cost']###所有成本存储3位小数，展示2位小数
            if former_cost != cost:
                self.update_cost(cost,name,supply)
                E_cost.delete(0, END)##理论上UI界面的操作应该在外部处理qaq（屎山代码的一部分）
                E_cost.insert(END, round(cost,2))
            if update_date and get_or_default(date,'') == res[3]:##date没有主动设置（与数据库保存值一致），则自动更新为今天
                self.execute("UPDATE valves SET date=? where name=?",(Today,name))
                self.conn.commit()
                E_date.delete(0, END)
                E_date.insert(END, Today)
            return cost

        def update_cost(self, cost, name, supply):
            self.execute("UPDATE re_supplies SET cost=? WHERE name=? AND supply=?", (cost,name,supply))
            self.conn.commit()

        def search_all(self):
            results=[]
            self.execute("""SELECT v.name, rs.supply, rs.cost, v.date, v.weight, v.hnumber, rs.bweight, rs.bwcost, rs.process, v.rodd, v.rodl, v.diff, v.remark FROM valves v JOIN re_supplies rs ON v.name = rs.name 
                        WHERE v.name LIKE '%'""")
            adds=[]#用于更新成本
            for r in self.cur.fetchall():
                rr = list(r)
                self.execute("SELECT items.iname,relationships.count,items.icost FROM relationships JOIN items on relationships.iname=items.iname where relationships.vname=?", (rr[0],))
                rr.append(list(self.cur.fetchall()))
                #计算成本，并记录，将更新成本一起修改
                cost_=round(sum(x[1]*x[2] for x in rr[-1])+rr[8]+rr[11]+cal_rod(rr[0],rr[9],rr[10],False)+(rr[7] if rr[6] == 0 else rr[6] * rr[7]),3)
                results.append(rr)
                if rr[2] != cost_:
                    rr[2]=cost_
                    adds.append((cost_,rr[0],rr[1]))
            self.executemany("UPDATE re_supplies SET cost=? WHERE name=? AND supply=?",adds)
            return results

    #补充函数
    def set_callback_value(value):
        """逆天回调函数，用全局变量传值，谨慎使用"""
        global value_callback
        value_callback = value

    def LB2PN(lb):
        """压力美标转国标"""
        return {
            "150": "25",
            "300": "50",
            "400": "64",
            "600": "100",
            "900": "150",
            "1500": "250",
            "2500": "420",
        }.get(lb, "0")

    def get_or_default(var, default=None):
        """获取输入框的值，如果为空则用default作为其默认值(混合了Stringvar和Entry)"""
        ori = var.get().strip()#删除前后空格
        if ori == '':
            if default is not None:
                return default
            else:
                return 0
        elif default is None:#对数字类型的，允许算术表达式
            try:
                num=eval(ori)
            except:
                return 0#数字类型却无法得到数字的，安全起见返回0
            else:
                return round(num,3)
        return ori
        
    def get_iname(default=None):
        """从iname部分的三个输入框取得实际的iname"""
        temp = iname_type.get().strip()
        itype = temp if temp != '名称' else ''
        temp = iname_spe.get().strip()
        ispe = temp if temp != '规格' else ''
        temp = iname_tex.get().strip()
        itex = temp if temp != '材质' else ''
        ans=f'{itype}-{ispe}-{itex}'
        if ans == '--' and default is not None:
            return default
        return '-'.join(ans.split('-')[:3])##永远只接受3个参数，变相阻止用户在iname中额外输入'-'

    def set_iname(iname):
        """根据iname的值设置iname部分的三个输入框"""
        itype,ispe,itex=iname.split('-')
        E_iname_type.delete(0, END)
        E_iname_type.add(itype)
        E_iname_type.event_generate('<FocusOut>')
        E_iname_spe.delete(0, END)
        E_iname_spe.add(ispe)
        E_iname_spe.event_generate('<FocusOut>')
        E_iname_tex.delete(0, END)
        E_iname_tex.add(itex)
        E_iname_tex.event_generate('<FocusOut>')

    def format_date(date_str):
        """将输入的更新日期转换为数据库中的标准格式"""
        match = re.match(r'(\d+)\D+(\d+)\D*(\d*)',date_str)
        if match:
            year,month,day = match.groups()
            if len(year) == 4:
                intyear = int(year)
            else:
                intyear = 2000 + int(year)
            if not day:
                day = '01'
            try:
                return datetime.datetime(intyear,int(month),int(day)).strftime('%Y-%m-%d')
            except ValueError:
                return Today
        return Today

    def cal_rod(name, rodd, rodl, check_flag=True):
        """计算阀杆价格"""
        duan = 0
        if '+' in rodl:###阀杆长度有加的才有锻
            duan = 1
        if name[0] == 'H':
            return 0
        try:
            rodd = float(rodd)
            rodl = float(eval(rodl))
        except (ValueError,SyntaxError):
            return 0
        duan *= float(config_duan.get(rodd, 0))
        namelist = name.split('-')
        if not namelist[1]:
            namelist[1]='10'##没有压力值的按低压处理（哈锅型）
        if namelist[2][-1] == '"':##口径修正
            namelist[2] = str(int(namelist[2][:-1])*25)
        if namelist[0].startswith('DK') or namelist[0].startswith('NK'):
            tap = namelist[0][2] + namelist[2]
        else:
            tap = namelist[0][0] + namelist[2]
        ###材质判断，先根据压力项后接字符判断，后根据后缀判断，互为补充，暂未考虑冲突情况
        vtype = 0   ##标记材质种类，0代表普通，1代表IV，2代表P(304/321)，3代表L(316L)
        if namelist[1][0] == 'P':#去除压力项中的温度参数
            namelist[1] = namelist[1][3:]
        temp = ''
        for i, char in enumerate(namelist[1]):
            if not char.isdigit():
                namelist[1] = namelist[1][:i]
                temp = namelist[1][i:]
        if temp.startswith('LB'):
            namelist[1] = LB2PN(namelist[1])    #美标压力修正
            temp=temp[2:]
        if temp == 'I' or temp == 'V':
            vtype = 1
        elif temp == 'P':#P是304材质
            vtype = 2
        elif temp == 'L':
            vtype = 3
        if len(namelist) > 3:#根据后缀补充判断材质
            if namelist[3] == 'WC6':
                vtype = 1
            elif namelist[3] == '316L':
                vtype = 3
            elif namelist[3] in ['304','321','CF8']:
                vtype = 2
        sorted_keys = list(config_fei.keys())
        index=-1
        try:
            index = sorted_keys.index(tap)
        except ValueError:
            #划分到最近档，强行给出近似结果
            for i in [25,]:#阀杆补丁
                try:
                    index=sorted_keys.index(tap[0]+str(int(tap[1:])+i))
                except ValueError:
                    continue
                else:
                    ntap=sorted_keys[index]
                    break
            if index == -1:
                if check_flag:
                    messagebox.showwarning("警告",f"{tap}飞不存在，请完善config.xlsx！")
                return 0
            else:
                if check_flag:
                    messagebox.showinfo("提示",f"{tap}飞不存在，已按{ntap}近似。")
                tap=ntap
        pressure = int(namelist[1])
        offset = 0
        if pressure >= 320:
            offset = 3
        elif pressure >= 160:
            offset = 2
        elif pressure >= 64:
            offset = 1
        if offset > 0:
            showdic['fei2'] = f'升{offset}档'
        else:
            showdic['fei2']=''
        index += offset
        try:##限位，分别对J、Z
            sorted_keys[index]
        except IndexError:
            if check_flag:
                messagebox.showwarning("警告",f"截止阀{tap[1:]}口径加{offset}档飞的价格未给出，请修改config.xlsx！")
            return 0
        if tap[0] == 'Z' and index > config_fei_limit:
            if check_flag:
                messagebox.showwarning("警告",f"闸阀{tap[1:]}口径加{offset}档飞的价格未给出，请修改config.xlsx！")
            return 0
        fei = float(config_fei.get(sorted_keys[index], 0))
        if vtype:
            fei = fei * 1.25
            showdic['fei2'] = showdic['fei2']+'特'
        if showdic['fei2'] != '':
            showdic['fei2'] = '('+showdic['fei2']+')'
        weight = rodd*rodd*0.0125*rodl/1000.0
        showdic['rodcost'] = config_rodcost[vtype]
        showdic['fei'] = fei
        showdic['duan'] = duan
        return weight*showdic['rodcost']+showdic['fei']+showdic['duan']

    def write_to_excel(wb):
        """将Workbook格式数据写回excel"""
        try:
            global observer
            observer.stop()
            wb.save(config_path)
        except PermissionError:
            messagebox.showerror("错误","请先关闭config.xlsx，或者直接在config.xlsx中修改。")
        finally:
            observer=Observer()
            observer.schedule(event_handler, path=config_path_, recursive=False)
            observer.start()

    def set_config():
        """将config.xlsx文件中的配置读取到内存中"""
        global config_fei
        global config_duan
        global config_fei_limit
        global config_rodcost
        global config_bwcost
        global config_texlist
        global config_typeslist
        config_fei = read_excel(config_path, sheet_name='飞', engine='openpyxl', usecols=[0,1]).set_index('口径')['价格'].to_dict()
        for index, key in enumerate(config_fei.keys()):
            if key[0] == 'J':
                config_fei_limit = index - 1
                break
        ###config_fei的第一列是字符，config_duan的第一列是纯数字，注意区别！
        config_duan = read_excel(config_path, sheet_name='锻', engine='openpyxl', usecols=[0,1]).set_index('阀杆外圆')['价格'].to_dict()
        df = read_excel(config_path, sheet_name='斤价', engine='openpyxl', header=None, usecols=[0,1])
        config_rodcost = [float(df.iloc[0,1]),float(df.iloc[1,1]),float(df.iloc[2,1]),float(df.iloc[3,1])]
        new_config_bwcost = read_excel(config_path, sheet_name='毛坯', engine='openpyxl', usecols=[0,1]).set_index('货源')['毛坯单价'].to_dict()
        if config_bwcost != new_config_bwcost:  #有更新时才刷新毛坯价格。# if 'config_bwcost' not in globals() or config_bwcost != new_config_bwcost:
            config_bwcost=new_config_bwcost.copy()
            window.event_generate("<<update_supplies>>") #通过自定义事件触发，使得update_supplies()在主进程中被调用。参数通过全局变量config_bwcost传递。
        config_texlist = read_excel(config_path, sheet_name='材质', engine='openpyxl', usecols=[0], dtype=str)['材质'].dropna().tolist()
        config_typeslist_ = read_excel(config_path, sheet_name='材质', engine='openpyxl', usecols=[2], dtype=str)['散件'].dropna().tolist()
        config_typeslist = [''] + config_typeslist_


    #触发逻辑函数
    def get_selected_row(event):
        """将item_list(Treeview)中选中的项对应显示到右侧的配件名称，配件价格，配件数量中"""
        if selection:=item_list.selection():    #如果有选中才处理，为了防止取消选中时触发该函数报错
            iid=selection[0]
            selected_list = item_list.item(iid,"values")
            cache['tv_selected'] = item_list.get_children().index(iid)
            E_iname_type.delete(0, END)
            E_iname_type.add(selected_list[0])
            E_iname_spe.delete(0, END)
            E_iname_spe.add(selected_list[1])
            E_iname_tex.delete(0, END)
            E_iname_tex.add(selected_list[2])
            E_icost.delete(0, END)
            E_icost.insert(END, selected_list[4])
            E_icount.delete(0, END)
            E_icount.insert(END, selected_list[3])

    def iname_compare(x, y):
        """配件名称大小比较，用于指定自动显示的配件的顺序"""
        x_=x.split('-')
        x1 = ''.join(x_[:2])
        x2 = x_[2]
        y_=y.split('-')
        y1 = ''.join(y_[:2])
        y2 = y_[2]
        ###先比较名称-规格的长度
        if len(x1) > len(y1):
            return 1
        elif len(x1) < len(y1):
            return -1
        ###再比较名称-规格的字典序
        if x1 > y1:
            return 1
        elif x1 < y1:
            return -1
        ###然后比较材质长短(空在前)
        if len(x2) > len(y2):
            return 1
        elif len(x2) < len(y2):
            return -1
        ###最后比较材质（逆序，即中碳-B7-316L-304）
        if x2 > y2:
            return -1
        elif x2 < y2:
            return 1
        return 0

    def autocomplete(event):
        """根据型号或配件名称的当前输入，匹配所有满足前缀的阀门或配件并显示"""
        global autocomplete_toplevel,autocomplete_listbox
        # 如果已经有一个autocomplete_toplevel存在，则销毁它
        if 'autocomplete_toplevel' in globals():#严格应该先判断in globals()再调用winfo_exists()##只用autocomplete_toplevel.winfo_exists()检查的话，在第一次声明前会报错，因为不知道autocomplete_toplevel是一个Toplevel。只用in globals()判断虽然不严谨，但是很好解决了第一次声明前的问题，并且即使重复调用destroy也不会有问题。)
            autocomplete_toplevel.destroy()
        if event.keysym == 'Return':
            return
        if event.widget is E_name:
            x_, y_ = E_name.winfo_rootx(), E_name.winfo_rooty()+E_name.winfo_height()
            width_ = E_name.winfo_width()
            if not (prefix := name.get()):
                return
            if not (matches := db.fsearch_valvename(prefix)):
                return
            matches.sort()
        else:   #if event.widget in (E_iname_type,E_iname_spe,E_iname_tex)
            x_, y_ = E_iname_type.winfo_rootx(), E_iname_type.winfo_rooty()+E_iname_type.winfo_height()
            width_ = E_iname_type.winfo_width()+E_iname_spe.winfo_width()+E_iname_tex.winfo_width()
            if not (prefix := get_iname('')):
                return
            if not (matches := db.fsearch_itemname(prefix)):
                return
            matches.sort(key=cmp_to_key(iname_compare))##满足短前缀在前，且中碳-B7-316L的顺序

        autocomplete_toplevel = Toplevel(window)
        autocomplete_toplevel.overrideredirect(True)  #移除窗口的边框和标题栏
        character_width = width_//8 #将像素宽度粗略转换成字符数
        autocomplete_toplevel.geometry("+%d+%d" % (x_,y_))
        autocomplete_listbox = Listbox(autocomplete_toplevel, width=character_width, height=min(len(matches),3), font=efont)
        autocomplete_listbox.pack(fill=BOTH, expand=True)
        for match in matches:
            autocomplete_listbox.insert(END, match)

        if event.keysym == 'Down':
            autocomplete_listbox.selection_set(0)
            autocomplete_listbox.focus_set()

        # 为Listbox绑定选择事件
        autocomplete_listbox.bind("<ButtonRelease-1>", lambda e:autocomplete_select(e.widget,event.widget))
        autocomplete_listbox.bind("<Up>", lambda e:list_up(e,event.widget))
        autocomplete_listbox.bind("<Down>", lambda e:(e.widget.selection_set(e.widget.curselection()[0]+1),e.widget.see(e.widget.curselection()[0]+1)))
        autocomplete_listbox.bind("<Return>", lambda e:autocomplete_select(e.widget,event.widget))
        autocomplete_listbox.bind("<KP_Enter>", lambda e:autocomplete_select(e.widget,event.widget))

    def autocomplete_select(widget,entry,selection=None):
        """从满足前缀的匹配结果中选中，填入到对应输入框中并执行查询"""
        if selection == None:
            selection = widget.get(widget.curselection()[0])
        if entry is E_name:
            entry.delete(0, END)
            entry.insert(0, selection)
            search_command(cursupply='')
        else:   #if event.widget in (E_iname_type,E_iname_spe,E_iname_tex)
            set_iname(selection)
            E_icost.delete(0, END)
            E_icost.insert(0, db.search_item(selection)[1])
            E_icost.focus_set()
            E_icost.selection_range(0,END)
        autocomplete_toplevel.destroy()
        window.clipboard_clear()  
        window.clipboard_append(selection)

    def autocomplete_cancel(event):
        """点击其他地方时自动取消自动匹配框"""
        if 'autocomplete_toplevel' in globals():
            if event.widget in [E_name,E_iname_type,E_iname_spe,E_iname_tex] and event.keysym != 'Tab':
                return
            autocomplete_toplevel.destroy()
        if 'supply_toplevel' in globals():
            if event.widget in [E_supply]:
                return
            supply_toplevel.destroy()

    def list_up(event,entry):
        """在自动匹配框中支持键盘↑键的操作"""
        index = event.widget.curselection()[0]
        if index == 0:
            entry.focus_set()
        else:
            index -= 1
            event.widget.selection_set(index)
            event.widget.see(index)

    def show_supply():
        """展示当前型号的所有货源"""
        global supply_toplevel
        if 'supply_toplevel' in globals():
            supply_toplevel.destroy()

        options=db.search_supplies(get_or_default(name,''))
        if not options:
            return
        supply_toplevel = Toplevel(window)
        supply_toplevel.overrideredirect(True)  #移除窗口的边框和标题栏
        character_width = E_supply.winfo_width()//8 #将像素宽度粗略转换成字符数
        supply_toplevel.geometry("+%d+%d" % (E_supply.winfo_rootx(), E_supply.winfo_rooty() + E_supply.winfo_height()))
        supply_listbox = Listbox(supply_toplevel, width=character_width, height=min(len(options)+1,3), font=efont)
        supply_listbox.pack(fill=BOTH, expand=True)
        supply_listbox.bind("<ButtonRelease-1>", lambda e:supply_select(e,options))
        for option in options:
            supply_listbox.insert(END, option[1])
        supply_listbox.insert(END, '')

    def supply_select(event,options):
        """从展示的所有货源中选中，并展示相应部分"""
        index=event.widget.curselection()[0]
        try:
            option=options[index]
        except IndexError:
            cost_=showdic['cost']-showdic['bw']-showdic['process']
            option=('','',cost_,'','','')
        ##不存在变量名冲突，可直接通过set函数设置
        supply.set(option[1])
        E_supply.focus_set()
        cost.set(round(option[2],2))
        bweight.set(option[3])
        bwcost.set(option[4])
        process.set(option[5])
        # 销毁下拉列表窗口
        supply_toplevel.destroy()
        window.event_generate('<KeyPress-Right>', keysym='Right')#模拟方向右键以显示光标

    def E_remark_tab():
        """为备注Text控件单独处理Tab输入，执行跳转逻辑而禁止输入Tab"""
        b2.focus_set()
        return "break"

    def E_remark_stab():
        E_diff.focus_set()
        E_diff.selection_range(0,END)
        return "break"

    def clear_treeview(treeview):
        """清空配件明细Treeview"""
        for item in treeview.get_children():
            treeview.delete(item)

    def refresh(name):
        """更新配件明细Treeview"""
        clear_treeview(item_list)
        for item in db.search_items(name):
            values=(item[0],item[1],item[2],item[3],item[4],round(item[3]*item[4],2))
            item_list.insert("", "end", f"{item[0]}-{item[1]}-{item[2]}", values=values)

    def show(res=[None,'未指定','',Today,0.0,'',0.0,0.0,0.0,'','',0.0,'']):
        """展示当前型号的查询结果"""
        if res is None:
            res=['','','','','','','','','','','','','']
        clear_treeview(item_list)
        showdic['name'],showdic['supply'],showdic['process'],showdic['rodd'],showdic['diff'] = res[0],res[1],res[8],res[9],res[11]
        try:
            showdic['rodl'] = eval(res[10])
        except SyntaxError:
            showdic['rodl'] = ''
        former_cost,cost = res[2],0
        if res[0]:#同时执行db.cal_cost_和refresh函数，减少查询数据库次数
            for item in db.search_items(res[0]):
                cost_ = item[3] * item[4]
                values=(item[0],item[1],item[2],item[3],item[4],round(cost_,2))
                item_list.insert("", "end", f"{item[0]}-{item[1]}-{item[2]}", values=values)
                cost += cost_
            showdic['items'] = cost
            showdic['rod'] = cal_rod(res[0],res[9],res[10],True)
            showdic['bw'] = res[7] if res[6]==0 else res[6] * res[7]
            showdic['cost'] = round(showdic['bw'] + showdic['process'] + showdic['rod'] + showdic['diff'] + showdic['items'],3)
            cost = showdic['cost']###所有成本存储3位小数，展示2位小数
            if former_cost != cost:
                db.update_cost(cost,res[0],res[1])
        E_name.delete(0, END)
        E_name.insert(END, res[0])
        E_supply.delete(0, END)
        E_supply.insert(END, res[1])
        E_cost.delete(0, END)
        E_cost.insert(END, round(cost,2))
        E_date.delete(0, END)
        E_date.insert(END, res[3])
        E_weight.delete(0, END)
        E_weight.insert(END, res[4])
        E_hnumber.delete(0, END)
        E_hnumber.insert(END, res[5])
        E_bweight.delete(0, END)
        E_bweight.insert(END, res[6])
        E_bwcost.delete(0, END)
        E_bwcost.insert(END, res[7])
        E_process.delete(0, END)
        E_process.insert(END, res[8])
        E_rodd.delete(0, END)
        E_rodd.insert(END, res[9])
        E_rodl.delete(0, END)
        E_rodl.insert(END, res[10])
        E_diff.delete(0, END)
        E_diff.insert(END, res[11])
        E_remark.delete('1.0', END)
        E_remark.insert(END, res[12])
        E_iname_type.clear()
        E_iname_spe.clear()
        E_iname_tex.clear()
        E_icost.delete(0, END)
        E_icount.delete(0, END)

    def on_closing():
        """安全关闭程序"""
        if messagebox.askokcancel("退出", "确认退出?\n请注意保存修改的备注。"):
            global observer
            observer.stop()
            observer.join()
            db.conn.close()
            window.destroy()

    def iname_direct(event, entry, bool):
        """为iname部分的三个输入框添加左右方向键逻辑"""
        if bool:
            if event.widget.index(INSERT) == 0:
                entry.focus_set()  
                entry.icursor(END)
        else:
            if event.widget.index(INSERT) == event.widget.index(END):
                entry.focus_set()  
                entry.icursor(0)

    def left2tv(event):
        """配件部分(和treeview)左方向键逻辑：跳转到treeview中"""
        if event.widget is not item_list and event.widget.index(INSERT) != 0:
            return
        if 'autocomplete_toplevel' in globals():
            autocomplete_toplevel.destroy()
        selected=item_list.selection()
        if not selected:#先前没有选中
            if not (children := item_list.get_children()):#treeview内容为空
                return
            if cache['tv_selected'] is None:
                selected = children[0]
            else:
                try:
                    selected=children[cache['tv_selected']]
                except:
                    selected=children[-1]
        item_list.focus_set()
        item_list.selection_set(selected)
        item_list.focus(selected)
        item_list.see(selected)

    def tv_right(event):
        """treeview右键逻辑：取消选中并将光标移动到配件价格"""
        selected=item_list.selection()
        if selected:
            item_list.selection_remove(item_list.selection()[0])    #仍会自动触发<<treeviewselect>>，因为selection发生改变。
            E_icost.focus_set()
            E_icost.selection_range(0,END)
            E_icost.icursor(0)
        else:
            E_iname_type.focus_set()
        return "break"  #禁用按右键触发的<<treeviewselect>>传播

    def tv_delete(event):
        selected=item_list.selection()
        if selected and get_iname('') == selected[0]:
            delitem_command()

    #按钮触发函数
    def select_valve():
        """型号按钮：直接选取型号以替换键盘输入"""
        def choose():
            name_list=[name1.get(),name2.get(),name3.get()]
            if name4.get() != '':
                name_list.append(name4.get())
                if name5.get() != '':
                    name_list.append(name5.get())
            valves_toplevel.destroy()
            global name
            name.set('-'.join(name_list))
            search_command(cursupply='')
            E_name.focus_set()
            E_name.icursor(END)
            window.event_generate('<KeyPress-Right>', keysym='Right')#模拟方向右键以显示光标

        def extract_first_number(s):  
            number_str = ''  
            for char in s:  
                if char.isdigit():  
                    number_str += char  
                elif number_str:  
                    break  
            try:  
                return int(number_str)  
            except ValueError:   
                return 0
        
        def update_comboboxes(event):
            name_list=[get_or_default(name1,'%'),get_or_default(name2,'%'),get_or_default(name3,'%')]
            len=3
            if name4.get() != '':
                name_list.append(name4.get())
                len=4
            if name5.get() != '':
                if len == 3:
                    name_list.append('%')
                name_list.append(name5.get())
                len=5
            if event.widget is name1:
                name_list[0]='%'
            elif event.widget is name2:
                name_list[1]='%'
            elif event.widget is name3:
                name_list[2]='%'
            elif event.widget is name4 and len >= 4:
                name_list[3]='%'
            elif event.widget is name5 and len >= 5:
                name_list[4]='%'

            results=db.fsearch_valvename('-'.join(name_list))
            __set(results)

        def __set(results):
            results.append('----')
            namevalues = {i: [] for i in range(1, 6)}
            for res in results:
                parts = res.split('-')
                for i,part in enumerate(parts):
                    namevalues[i+1].append(part)
            namevalues[1] = sorted(set(namevalues[1]))
            for i in range(2,4):
                namevalues[i] = list(set(namevalues[i]))
                namevalues[i].sort(key=extract_first_number)
            namevalues[4] = sorted(set(namevalues[4]))
            namevalues[5] = sorted(set(namevalues[5]))
            name1['values'] = namevalues[1]
            name2['values'] = namevalues[2]
            name3['values'] = namevalues[3]
            name4['values'] = namevalues[4]
            name5['values'] = namevalues[5]

        global valves_toplevel
        if 'valves_toplevel' in globals():
            valves_toplevel.destroy()
        valves_toplevel = Toplevel(window)
        valves_toplevel.title('选择型号')
        valves_toplevel.geometry("+%d+%d" % (l_name.winfo_rootx(), l_name.winfo_rooty()+l_name.winfo_height()))
        label = Button(valves_toplevel, font=lfont, width=6, text="选择型号")
        label.grid(row=0, column=0, padx=35, pady=10)
        label.configure(command=choose)
        name1 = ttk.Combobox(valves_toplevel, width=7)
        name1.grid(row=0, column=1)
        name1.bind("<Button-1>", update_comboboxes)
        label1 = Label(valves_toplevel, font=lfont, width=1, text="-")
        label1.grid(row=0, column=2)
        name2 = ttk.Combobox(valves_toplevel, width=5)
        name2.grid(row=0, column=3)
        name2.bind("<Button-1>", update_comboboxes)
        label2 = Label(valves_toplevel, font=lfont, width=1, text="-")
        label2.grid(row=0, column=4)
        name3 = ttk.Combobox(valves_toplevel, width=5)
        name3.grid(row=0, column=5)
        name3.bind("<Button-1>", update_comboboxes)
        label3 = Label(valves_toplevel, font=lfont, width=1, text="-")
        label3.grid(row=0, column=6)
        name4 = ttk.Combobox(valves_toplevel, width=5)
        name4.grid(row=0, column=7)
        name4.bind("<Button-1>", update_comboboxes)
        label4 = Label(valves_toplevel, font=lfont, width=1, text="-")
        label4.grid(row=0, column=8)
        name5 = ttk.Combobox(valves_toplevel, width=5)
        name5.grid(row=0, column=9)
        name5.bind("<Button-1>", update_comboboxes)
        valves_toplevel.columnconfigure(10,minsize=30)
        results = db.fsearch_valvename('*')
        __set(results)

    def show_compare():
        """货源按钮：提供如同excel的编辑操作"""
        curname=get_or_default(name,'')
        supplies = db.search_supplies(curname)
        if not supplies:
            return
        compare_toplevel = Toplevel(window)
        compare_toplevel.title('货源详情')
        compare_toplevel.geometry("+%d+%d" % (window.winfo_rootx(), window.winfo_rooty()))
        compare_toplevel.transient(window)
        compare_toplevel.grab_set()
        compare_toplevel.focus_set()
        table = SimpleTableInput(compare_toplevel, supplies, showdic['cost']-showdic['bw']-showdic['process'])
        table.pack(anchor='center')
        compare_toplevel.protocol("WM_DELETE_WINDOW", lambda table=table:compare_close(table,compare_toplevel,supplies))
        window.wait_window(compare_toplevel)

    def compare_close(table,compare_toplevel,old_supplies):
        """货源按钮拓展：关闭时自动更新货源"""
        old_supplies_listlist = [list((tup[0],tup[1],round(tup[2],2),tup[3],tup[4],tup[5])) for tup in old_supplies]
        new_supplies,othercost,name = table.get()
        if new_supplies == old_supplies_listlist:##未修改直接返回
            compare_toplevel.grab_release()
            compare_toplevel.destroy()
            return
        if messagebox.askokcancel("提示",f"是否要修改保存？"):
            for new_supply in new_supplies:
                if new_supply[3] == 0:
                    new_supply[2] = round(new_supply[4] + new_supply[5] + othercost,3)
                else:
                    new_supply[2] = round(new_supply[3]*new_supply[4] + new_supply[5] + othercost,3)
            if db.reset_supplies(new_supplies,name) or messagebox.askokcancel("提示",f"货源修改失败，是否仍要退出？"):
                compare_toplevel.grab_release()
                compare_toplevel.destroy()
                search_command(cursupply='')
                return
        else:
            compare_toplevel.grab_release()
            compare_toplevel.destroy()
                    
    def del_supply():
        """货源删除按钮"""
        curname=name.get()
        cursupply=supply.get()
        if messagebox.askokcancel("提示",f"确认删除'{curname}'的货源'{cursupply}'？"):
            db.delete_supply(curname,cursupply)
            E_supply.delete(0, END)
            search_command()

    ##货源下拉列表逻辑与自动填充框一致，写在触发逻辑函数中

    def show_costdetails():
        """成本按钮：显示当前的成本计算详情"""
        global showdic
        if showdic['name'] == None:
            showdic={'name':'','cost':'','bw':'','process':'','rod':'','diff':'','items':'','rodd':'','rodl':'','rodcost':'','fei':'','fei2':'','duan':''}
        elif showdic['name'] == '':
            messagebox.showinfo("成本计算详情",f"成本=毛坯费+加工费+阀杆+其他费用+配件总价\n其中阀杆=外圆*外圆*0.0125*总长/1000*斤价+飞+锻")
        elif showdic['name'][0] == 'H':    
            messagebox.showinfo("成本计算详情",f"{showdic['name']}{showdic['supply']}成本{round(showdic['cost'],2)}=毛坯费{round(showdic['bw'],2)}+加工费{round(showdic['process'],2)}+其他费用{round(showdic['diff'],2)}+配件总价{round(showdic['items'],2)}")
        else:
            messagebox.showinfo("成本计算详情",f"{showdic['name']}{showdic['supply']}成本{round(showdic['cost'],2)}=毛坯费{round(showdic['bw'],2)}+加工费{round(showdic['process'],2)}+阀杆{round(showdic['rod'],2)}+其他费用{round(showdic['diff'],2)}+配件总价{round(showdic['items'],2)}\n其中阀杆=外圆{showdic['rodd']}*外圆{showdic['rodd']}*0.0125*总长{showdic['rodl']}/1000*斤价{showdic['rodcost']}+飞{showdic['fei']}{showdic['fei2']}+锻{showdic['duan']}")

    def search_command(curname=None,cursupply=None):
        """查询按钮：按照当前的型号和货源查询"""
        cache['tv_selected'] = None
        if curname is None:
            curname=get_or_default(name,'')
        if cursupply is None:
            cursupply=get_or_default(supply,'')
        if curname == '':
            show(None)
            return
        res = db.search_valve(curname,cursupply)
        if res == None:
            if messagebox.askokcancel("提示",f"当前型号'{curname}'未创建，是否新建？"):
                db.insert_valve(curname)
                show(db.search_valve(curname))
        else:
            show(res)

    def update_command():
        """修改按钮：将当前值保存到数据库中"""
        curname=get_or_default(name,'')
        cursupply=get_or_default(supply,'未指定')
        curdate=format_date(get_or_default(date,Today))
        if curname == '':
            messagebox.showinfo("提示","型号不能为空！")
            return
        res = db.search_valve(curname)
        if res == None:
            if messagebox.askokcancel("提示","当前型号未创建，是否按当前参数创建？"):
                db.insert_valve(curname, cursupply, get_or_default(cost), curdate, get_or_default(weight), get_or_default(hnumber,''), get_or_default(bweight), get_or_default(bwcost), get_or_default(process), get_or_default(rodd,''), get_or_default(rodl,''), get_or_default(diff), E_remark.get(1.0,END).strip())
                ##把当前配件也一并添加（仅针对使用'修改'按钮新建型号的情况）
                relationships=[]
                for row in item_list.get_children():  
                    item = item_list.item(row, 'values')
                    iname=f'{item[0]}-{item[1]}-{item[2]}'
                    relationships.append((curname,iname,item[3]))
                db.insert_relationships_nocheck(relationships)
                search_command(curname,cursupply)
        else:
            if messagebox.askokcancel("提示",f"确认修改型号'{curname}'货源'{cursupply}'的阀门参数？"):
                if get_or_default(date,'') == res[3]:##原始输入日期等于数据库日期，才认为没有手动设置日期
                    curdate = Today
                db.update_valve(curname, cursupply, get_or_default(cost), curdate, get_or_default(weight), get_or_default(hnumber,''), get_or_default(bweight), get_or_default(bwcost), get_or_default(process), get_or_default(rodd,''), get_or_default(rodl,''), get_or_default(diff), E_remark.get(1.0,END).strip())
                search_command(curname,cursupply)

    def sort_tree(tv, col, reverse=False):  
        """对Treeview控件内容排序，其中数量分整型和浮点型分别排序"""
        # 获取列名对应的索引
        col_index = tv['columns'].index(col)
        # 获取所有项及其值
        l = [(tv.item(k, 'values')[col_index], k) for k in tv.get_children('')]
        # 排序
        if col == '数量':##数量将整型数和浮点型数分别排序（对应的是两类配件，斤和个数的（若有整数斤会出错，不过对于排序来说问题不大））
            l1=[]
            l2=[]
            for t in l:
                if float(t[0]).is_integer():
                    l1.append(t)
                else:
                    l2.append(t)
            l1.sort(reverse=reverse, key=lambda t: float(t[0]))
            l2.sort(reverse=reverse, key=lambda t: float(t[0]))
            l=l2+l1
        else:
            l.sort(reverse=reverse, key=lambda t: float(t[0]))
        # 重新排列项
        for index, (_, k) in enumerate(l):
            tv.move(k, '', index)
        # 更新命令以切换排序顺序
        tv.heading(col, text=col, command=lambda col=col: sort_tree(tv, col, not reverse))

    def additem_command():
        """添加配件按钮：如果未输入数量，则仅添加或修改配件；如果未输入价格，则读取数据库中的值或记为0，并添加到当前型号"""
        curname=get_or_default(name,'')
        curiname=get_iname('')
        curicost=get_or_default(icost)
        curcount=get_or_default(icount)
        if curiname == '':
            messagebox.showinfo("提示","配件名不能为空！")
            E_iname_type.focus_set()
            E_iname_type.selection_range(0,END)
            return
        res = db.search_item(curiname)
        if res is None:
            window.wait_window(ChoiceDialog(window,curname,curiname,curicost,curcount))
            E_iname_type.focus_set()
            E_iname_type.selection_range(0,END)
            return
        if curcount == 0:
            if messagebox.askokcancel("提示","未输入配件数量，是否仅修改配件价格？"):
                db.insert_item(curiname,curicost)
                set_iname(curiname)
                E_icost.delete(0, END)
                E_icost.insert(END, curicost)
                show(db.search_valve(curname))
            E_icost.focus_set()
            E_icost.selection_range(0,END)
            return
        if curicost == 0 or float(curicost) == res[1]:
            if curname == '':
                messagebox.showinfo("提示","型号名为空，添加失败！")
                E_name.focus_set()
                return
            db.insert_relationship(curname,curiname,curcount)
            set_iname(curiname)
            E_icost.delete(0, END)
            E_icost.insert(END, curicost)
            E_icount.delete(0, END)
            E_icount.insert(END, curcount)
            refresh(curname)
            E_iname_type.focus_set()
            E_iname_type.selection_range(0,END)
            cache['tv_selected']=item_list.get_children().index(curiname)
            return
        if messagebox.askokcancel("提示","当前配件价格发生变化，是否修改并添加？"):
            db.insert_item(curiname,curicost)
            set_iname(curiname)
            E_icost.delete(0, END)
            E_icost.insert(END, curicost)
            if curname == '':
                messagebox.showinfo("提示","型号名为空添加失败，仅修改配件价格！")
                refresh(curname)
                E_icost.focus_set()
                E_icost.selection_range(0,END)
                return
            if curcount == 0:
                db.cal_cost_(curname)
                refresh(curname)
                messagebox.showinfo("提示","配件数量为空，仅修改配件价格！")
                E_icost.focus_set()
                E_icost.selection_range(0,END)
            else:
                db.insert_relationship(curname,curiname,curcount)
                E_icount.delete(0, END)
                E_icount.insert(END, curcount)
                E_iname_type.focus_set()
                E_iname_type.selection_range(0,END)
                refresh(curname)
                cache['tv_selected']=item_list.get_children().index(curiname)

    def delitem_command():
        """删除配件按钮：将配件从当前型号中删除"""
        curname=get_or_default(name,'')
        curiname=get_iname('')
        curcount=get_or_default(icount)
        if curiname == '':
            messagebox.showinfo("提示","配件名为空，无法删除！")
            E_iname_type.focus_set()
            return
        if db.delete_relationship(curname,curiname,curcount):
            E_iname_type.clear()
            E_iname_spe.clear()
            E_iname_tex.clear()
            E_icost.delete(0, END)
            E_icount.delete(0, END)
            refresh(curname)
            item_list.focus_set()

    def checkitem_mcommand():
        """菜单栏查看配件：方便查看配件的一些操作"""
        def select_items():
            clear_treeview(items_list)
            for res in db.fsearch_item_valves(finame.get()):
                items_list.insert("", "end", values=res)    
        
        def items_list_select(event):
            selected_list = items_list.item(items_list.selection(),"values")
            window.clipboard_clear()  
            window.clipboard_append(selected_list[2])

        global checkitem_toplevel
        if 'checkitem_toplevel' in globals():
            checkitem_toplevel.destroy()
        checkitem_toplevel = Toplevel(window,padx=30,pady=10)
        checkitem_toplevel.title('查看配件')
        width,height = 500,400
        screen_width,screen_height = checkitem_toplevel.winfo_screenwidth(),checkitem_toplevel.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        checkitem_toplevel.geometry(f"{width}x{height}+{x}+{y}")
        checkitem_toplevel.grid_rowconfigure(1, weight=1)
        checkitem_toplevel.grid_columnconfigure(0, weight=1)
        checkitem_toplevel.grid_columnconfigure(3, weight=1)
        finame=StringVar()
        E_finame=SuperEntry(checkitem_toplevel, placeholder='请输入要搜索的配件名', font=efont, width=20, textvariable=finame)
        E_finame.grid(row=0, column=1, padx=30, pady=15)
        B_finame=Button(checkitem_toplevel, font=lfont, width=6, text="模糊查询", command=select_items)
        B_finame.grid(row=0, column=2, padx=30, pady=15)
        B_finame.bind("<Return>", lambda e:select_items())
        B_finame.bind("<KP_Enter>", lambda e:select_items())
        items_list = ttk.Treeview(checkitem_toplevel, columns=("名称","单价","包含在下列型号中"), show="headings")
        items_list.heading("名称", text="名称")
        items_list.heading("单价", text="单价")
        items_list.heading("包含在下列型号中", text="包含在下列型号中")
        items_list.column("名称", width=120, stretch=NO)
        items_list.column("单价", width=50, stretch=NO)
        items_list.column("包含在下列型号中", width=300, stretch=YES)
        items_list.grid(row=1, column=0, columnspan=4, sticky='nsew')
        ct_ysb = Scrollbar(checkitem_toplevel)
        ct_ysb.grid(row=1, column=4, sticky='ns')
        ct_ysb.configure(command=items_list.yview, takefocus=False)
        items_list.configure(selectmode='browse', yscrollcommand=ct_ysb.set)
        items_list.bind('<<TreeviewSelect>>', items_list_select)
        checkitem_toplevel.focus_set()
        E_finame.focus_set()

    def delvalve_mcommand(vname, **kwargs):
        """菜单栏删除阀门（回调函数）：从数据库中删除输入的型号"""
        db.delete_valve(vname)
        if vname == get_or_default(name,str):
            E_name.delete(0, END)
            show(None)
        
    def delitem_mcommand(iname, **kwargs):
        """菜单栏删除配件（回调函数）：从数据库中删除输入配件"""
        db.delete_item(iname)
        search_command()

    def updrod_mcommand(cost,**kwargs):
        """菜单栏修改阀杆斤价（回调函数）：修改对应的阀杆斤价"""
        try:
            cost=float(cost)
        except ValueError:
            messagebox.showwarning("警告","请输入正确格式的阀杆斤价！")
        else:
            global config_rodcost
            config_rodcost[kwargs['vtype']] = cost
            wb = load_workbook(config_path)
            buf='B'+str(kwargs['vtype']+1)
            wb['斤价'][buf] = cost
            write_to_excel(wb)
            search_command()

    def input_mcommand():
        """菜单栏导入配件价格：从'配件价格.xlsx'导入配件价格'"""
        workbook = load_workbook(filename=items_path)
        items={}
        adds={}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for r in sheet.iter_rows(min_row=2, values_only=True):
                itype=str(r[0]) if r[0] else ''
                ispe=str(r[1]) if r[1] else ''
                itex=str(r[2]) if r[2] else ''
                try:  
                    icost = float(r[3]) if r[3] else 0  
                except:  
                    icost = 0
                if (iname := f'{itype}-{ispe}-{itex}') == '--':
                    continue
                if itype != '压子':#压子价格只受规格第一个参数影响，但是录入规格是有三个参数的，所以不将一个参数规格的压子录入数据库。
                    # items.append((iname,icost))
                    items[iname]=icost
                else:
                    spe=ispe.split('*')[0]
                    # adds.append((icost,f'{itype}-{spe}%-{itex}'))
                    adds[icost]=f'{itype}-{spe}%-{itex}'
                if itype == '石墨圈':
                    # adds.append((icost,itype+'-%-'+itex))
                    adds[icost]=itype+'-%-'+itex
                elif itype == '铜螺母':
                    # adds.append((icost,itype+'-%'))
                    adds[icost]=itype+'-%'
                elif itype[-2:] == '螺帽':
                    # adds.append((icost,f'%螺帽-{ispe}-{itex}'))
                    adds[icost]=f'%螺帽-{ispe}-{itex}'
        db.insert_items(tuple(items.items()),tuple(adds.items()))
        search_command()

    def output_mcommand():
        """菜单栏导出所有数据：将所有数据导出至'配件明细汇总_日期.xlsx'"""
        inout_menu.entryconfigure("导出所有数据", state='disabled')
        results=db.search_all()
        E_cost.delete(0, END)
        E_cost.insert(END, 0)
        output_queue=Queue()
        if not isfile('配件明细汇总导出.xlsx'):
            process = Process(target=output_p, args=(output_queue,results,template_path,order_H,order_J,order_Z))
        else:
            process = Process(target=output_p, args=(output_queue,results,'配件明细汇总导出.xlsx',order_H,order_J,order_Z))
        process.start()

        def check_process():
            if not process.is_alive():
                state = output_queue.get()
                if state:
                    messagebox.showinfo("成功", f"数据已成功导出到'配件明细汇总导出.xlsx'文件。")
                else:
                    messagebox.showinfo("提示", f"'配件明细汇总导出.xlsx'文件被占用，请关闭后再重新导出。")
                inout_menu.entryconfigure("导出所有数据", state='normal')
            else:
                window.after(1000, check_process)

        window.after(1000, check_process)  


    #新建窗体
    class inputDialog(Toplevel):
        """创建可以输入一个值的Toplevel，并在点击确定后调用相应的回调函数，将值传递给回调函数处理"""
        def __init__(self, parent, title, text, funcname, **kwargs):
            super().__init__(parent)
            self.title(title)
            self.geometry("250x170")
            self.center_window()
            label = Label(self, text=text)
            label.pack(pady=18)
            input = StringVar()
            E_input = Entry(self, textvariable=input)
            E_input.pack(pady=10)
            button_frame = Frame(self)
            button_frame.pack(pady=10)
            btn1 = Button(button_frame, text="确定", command=lambda:(funcname(input.get(),**kwargs),self.destroy()))
            btn1.bind("<Return>", lambda e:(funcname(input.get(),**kwargs),self.destroy()))
            btn1.bind("<KP_Enter>", lambda e:(funcname(input.get(),**kwargs),self.destroy()))
            btn2 = Button(button_frame, text="返回", command=self.destroy)
            btn2.bind("<Return>", lambda e:self.destroy())
            btn2.bind("<KP_Enter>", lambda e:self.destroy())
            btn1.pack(side="left",padx=30)
            btn2.pack(side="right",padx=30)
            self.transient(window)
            self.grab_set()
            self.focus_set()
            E_input.focus_set()
            self.protocol("WM_DELETE_WINDOW", lambda:(self.grab_release(),self.destroy()))

        def center_window(self):
            # 获取对话框的宽度和高度
            width = self.winfo_reqwidth()
            height = self.winfo_reqheight()
            # 获取屏幕宽度和高度
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
            # 计算x和y坐标使得窗口居中
            x = (screen_width // 2) - (width // 2)
            y = (screen_height // 2) - (height // 2)
            # 设置窗口的geometry属性
            self.geometry(f"+{x}+{y}")  # 注意这里不设置宽度和高度，因为它们已经由内容决定
        
    class ChoiceDialog(Toplevel):
        """添加配件时处理各种情况的对话框"""
        def __init__(self, parent, curname, curiname, curicost, curcount):
            super().__init__(parent)
            self.title("添加配件")
            self.geometry("350x200")
            self.center_window()

            label = Label(self, text="当前配件不存在，是否新建和添加至该型号？")
            label.pack(fill="x", pady=30)

            button_frame = Frame(self)
            button_frame.pack(expand=True, fill="x", pady=50)

            btn1 = Button(button_frame, text="仅新建配件", command=lambda:self.btn1_command(curiname, curicost))
            btn1.bind("<Return>", lambda e:self.btn1_command(curiname, curicost))
            btn1.bind("<KP_Enter>", lambda e:self.btn1_command(curiname, curicost))
            btn2 = Button(button_frame, text="新建并添加", command=lambda:self.btn2_command(curname, curiname, curicost, curcount))
            btn2.bind("<Return>", lambda e:self.btn2_command(curname, curiname, curicost, curcount))
            btn2.bind("<KP_Enter>", lambda e:self.btn2_command(curname, curiname, curicost, curcount))
            btn3 = Button(button_frame, text="取消并返回", command=self.destroy)
            btn3.bind("<Return>", lambda e:self.destroy())
            btn3.bind("<KP_Enter>", lambda e:self.destroy())
            btn1.pack(side="left", padx=20)
            btn2.pack(side="left", padx=25)
            btn3.pack(side="left", padx=20)

            self.transient(window)
            self.grab_set()
            self.focus_set()
            self.protocol("WM_DELETE_WINDOW", lambda:(self.grab_release(),self.destroy()))

        def btn1_command(self, curiname, curicost):
            db.insert_item(curiname, curicost)
            set_iname(curiname)
            E_icost.delete(0, END)
            E_icost.insert(END, curicost)
            self.destroy()

        def btn2_command(self, curname, curiname, curicost, curcount):
            db.insert_item_relationship(curname, curiname, curicost, curcount)
            set_iname(curiname)
            E_icost.delete(0, END)
            E_icost.insert(END, curicost)
            E_icount.delete(0, END)
            E_icount.insert(END, curcount)
            refresh(curname)
            try:
                cache['tv_selected']=item_list.get_children().index(curiname)
            except:
                pass
            self.destroy()
        
        def center_window(self):
            # 获取对话框的宽度和高度
            width = self.winfo_reqwidth()
            height = self.winfo_reqheight()
            # 获取屏幕宽度和高度
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
            # 计算x和y坐标使得窗口居中
            x = (screen_width // 2) - (width // 2)
            y = (screen_height // 2) - (height // 2)
            # 设置窗口的geometry属性
            self.geometry(f"+{x}+{y}")  # 注意这里不设置宽度和高度，因为它们已经由内容决定

    class ConfigWatcher(FileSystemEventHandler):
        """监测config.xlsx文件变化的类"""
        def on_modified(self, event):
            if event.src_path == config_path:
                set_config()
                window.event_generate("<<search_command>>") #通过自定义事件触发，使得search_command()在主进程中被调用。
                
    class SimpleTableInput(Frame):
        """一个简单的模拟excel表格的窗体，可以用get函数逐列取得数据"""
        def __init__(self, parent, options, othercost):
            Frame.__init__(self, parent)
            self._entry = {}
            self.columns = len(options) + 2
            self.rows = len(options[0]) + 1
            self.Button = None
            self.othercost = othercost
            # register a command to use for validation
            self.vcmd = (self.register(self._validate), "%P")
            # create the table of widgets
            column = 0
            text_=(options[0][0],'成本','毛坯重量','毛坯单价','加工费','其他费用')
            for row in range(self.rows-1):
                index = (column, row)
                e = Entry(self, width=12, font=efont)
                e.insert(END, text_[row])
                e.configure(state='readonly')
                e.grid(row=row, column=column, stick="nsew")
                self._entry[index] = e
            for column in range(1,self.columns-1):
                for row in range(0,self.rows-1):
                    index = (column, row)
                    if row == 0:
                        e = Entry(self, width=12, font=efont)
                        e.insert(END, options[column-1][row+1])
                    elif row == self.rows-2:
                        e = Entry(self, width=12, font=efont)
                        e.insert(END, round(self.othercost,2))
                        e.configure(state='readonly')
                    else:
                        e = Entry(self, width=12, font=efont, validate="key", validatecommand=self.vcmd)
                        e.bind("<KeyRelease>", lambda e,column=column:self.table_auto(column))
                        e.insert(END, options[column-1][row+1])
                        if row == 1:
                            value_=float(e.get())
                            e.delete(0, END)
                            e.insert(END, round(value_,2))
                            e.configure(state='readonly')
                    e.grid(row=row, column=column, stick="nsew")
                    self._entry[index] = e
                e = Button(self, text='－', width=2, command=lambda delcolumn=column:self.delcolumn(delcolumn), cursor='hand2', font=efont)
                e.grid(row=self.rows-1,column=column)
                self._entry[(column,self.rows-1)] = e
            # adjust column weights so they all expand equally
            for column in range(self.columns-1):
                self.grid_columnconfigure(column, weight=1)
            # designate a final, empty row to fill up any extra space
            self.grid_rowconfigure(self.rows, weight=1)
            f2=Frame(self)
            f2.grid(row=0, column=self.columns-1, rowspan=6)
            e = Button(f2, text='＋', command=self.addcolumn, cursor='hand2', font=efont)
            e.pack(anchor='center')
            self.Button_f=f2

        def get(self):
            '''Return a list of supplies'''
            name=self._entry[(0,0)].get()
            result = []
            for column in range(1,self.columns-1):
                current_column = []
                current_column.append(name)
                current_column.append(get_or_default(self._entry[(column,0)],'未指定'))
                for row in range(1,self.rows-2):
                    index = (column, row)
                    current_column.append(float(get_or_default(self._entry[index],0)))
                result.append(current_column)
            return result,self.othercost,name

        def addcolumn(self):
            self.columns += 1
            self.Button_f.grid(row=0, column=self.columns-1, rowspan=6)
            column = self.columns - 2
            for row in range(0,self.rows-1):
                index = (column, row)
                if row == 0:
                    e = Entry(self, width=12, font=efont)
                elif row == self.rows-2:
                    e = Entry(self, width=12, font=efont)
                    e.insert(END, round(self.othercost,2))
                    e.configure(state='readonly')
                else:
                    e = Entry(self, width=12, font=efont, validate="key", validatecommand=self.vcmd)
                    e.bind("<KeyRelease>", lambda e,column=column:self.table_auto(column))
                    if row == 1:
                        e.configure(state='readonly')
                e.grid(row=row, column=column, stick="nsew")
                self._entry[index] = e
            e = Button(self, text='－', width=4, command=lambda delcolumn=column:self.delcolumn(delcolumn), cursor='hand2', font=efont)
            e.grid(row=self.rows-1,column=column)
            self._entry[(column,self.rows-1)] = e

        def delcolumn(self, delcolumn):
            for row in range(self.rows-1):
                self._entry[(delcolumn,row)].destroy()
                del self._entry[(delcolumn,row)]
            for column in range(delcolumn,self.columns-2):
                for row in range(self.rows-1):
                    self._entry[(column,row)] = self._entry[(column+1,row)]
                    self._entry[(column,row)].grid(row=row,column=column)
            self._entry[(self.columns-2,self.rows-1)].destroy()
            if delcolumn != self.columns-2:
                for row in range(self.rows):
                    del self._entry[(self.columns-2,row)]
            self.Button_f.grid(row=0, column=self.columns-2, rowspan=6)
            self.columns -= 1

        def table_auto(self, column):
            cost = float(get_or_default(self._entry[(column,2)],1))*float(get_or_default(self._entry[(column,3)],0))+float(get_or_default(self._entry[(column,4)],0))+float(get_or_default(self._entry[(column,5)],0))
            self._entry[(column,1)].configure(state='normal')
            self._entry[(column,1)].delete(0, END)
            self._entry[(column,1)].insert(END, round(cost,2))
            self._entry[(column,1)].configure(state='readonly')

        def _validate(self, P):
            '''Perform input validation. 
            Allow only an empty value, or a value that can be converted to a float
            '''
            if P.strip() == "":
                return True

            try:
                f = float(P)
            except ValueError:
                self.bell()
                return False
            return True

    class SuperEntry(Entry):  
        def __init__(self, parent, placeholder, color='grey', cnf={}, **kw):
            super().__init__(parent, cnf, **kw)

            self.placeholder = placeholder
            self.placeholder_color = color
            self.default_fg_color = self["fg"]
            self.bind("<FocusIn>", self.foc_in)
            self.bind("<FocusOut>", self.foc_out)
            self.put_placeholder()

        def put_placeholder(self):
            self["fg"] = self.placeholder_color
            self.delete(0, END)
            self.insert(END, self.placeholder)

        def foc_in(self, *args):
            if self["fg"] == self.placeholder_color:
                self.delete(0, END)
                self["fg"] = self.default_fg_color

        def foc_out(self, *args):
            if not self.get():
                self.put_placeholder()
        
        def clear(self):
            self.delete(0, END)
            self.put_placeholder()

        def add(self, str):
            self["fg"] = self.default_fg_color
            self.insert(END, str)


    db = TTVALVE()
    config_bwcost=dict()##定义初值，减少set_config中的判断。
    showdic={'name':'','cost':'','bw':'','process':'','rod':'','diff':'','items':'','rodd':'','rodl':'','rodcost':'','fei':'','fei2':'','duan':''}
    # set_config()#由于函数中涉及窗口触发，在窗口创建后再调用
    order_H={}
    for idx, val in enumerate(tuple(read_excel(template_path, sheet_name='止回阀模板', skiprows=4).iloc[:, 0].dropna().to_list()[:-1])):
        order_H[val] = 2 * idx
        order_H[val+'成品'] = 2 * idx + 1
    order_J={}
    for idx, val in enumerate(tuple(read_excel(template_path, sheet_name='截止阀模板', skiprows=4).iloc[:, 0].dropna().to_list()[:-1])):
        order_J[val] = 2 * idx
        order_J[val+'成品'] = 2 * idx + 1
    order_Z={}
    for idx, val in enumerate(tuple(read_excel(template_path, sheet_name='闸阀模板', skiprows=4).iloc[:, 0].dropna().to_list()[:-1])):
        order_Z[val] = 2 * idx
        order_Z[val+'成品'] = 2 * idx + 1
    event_handler = ConfigWatcher()
    observer = Observer()
    observer.schedule(event_handler, path=config_path_, recursive=False)
    observer.start()
    cache={'tv_selected':None}

    #####界面设计
    window = Tk()
    window.title("天泰阀门")
    # window.resizable(False, False)
    default_font = nametofont("TkDefaultFont")
    default_font.config(size=10)
    window.option_add("*Font", default_font)
    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    window.protocol("WM_DELETE_WINDOW", on_closing)
    window.bind("<Configure>", autocomplete_cancel)
    window.bind("<Button-1>", autocomplete_cancel)
    window.bind("<<search_command>>",lambda e:search_command())
    window.bind("<<update_supplies>>",lambda e:db.update_supplies())
    set_config()


    menu_bar = Menu(window)
    menu_bar.config(font=default_font)
    file_menu = Menu(menu_bar, tearoff=0)
    file_menu.add_command(label="清理", command=db.delete_trash)
    file_menu.add_command(label="退出", command=on_closing)
    menu_bar.add_cascade(label="文件", menu=file_menu)
    check_menu = Menu(menu_bar, tearoff=0)
    # check_menu.add_command(label="查看阀门")
    check_menu.add_command(label="查看配件", command=checkitem_mcommand)
    menu_bar.add_cascade(label="查看", menu=check_menu)
    delete_menu = Menu(menu_bar, tearoff=0)
    delete_menu.add_command(label="删除阀门", command=lambda:window.wait_window(inputDialog(window,'删除阀门','请输入需要删除的阀门型号：',delvalve_mcommand)))
    delete_menu.add_command(label="删除配件", command=lambda:window.wait_window(inputDialog(window,'删除配件','请输入需要删除的配件名称：',delitem_mcommand)))
    menu_bar.add_cascade(label="删除", menu=delete_menu)
    update_menu = Menu(menu_bar, tearoff=0)
    update_menu.add_command(label="修改普通阀杆斤价", command=lambda:window.wait_window(inputDialog(window,'修改普通阀杆斤价','请输入新的普通阀杆斤价：',updrod_mcommand,vtype=0)))
    update_menu.add_command(label="修改IV阀杆斤价", command=lambda:window.wait_window(inputDialog(window,'修改IV阀杆斤价','请输入新的IV阀杆斤价：',updrod_mcommand,vtype=1)))
    update_menu.add_command(label="修改304阀杆斤价", command=lambda:window.wait_window(inputDialog(window,'修改304阀杆斤价','请输入新的304阀杆斤价：',updrod_mcommand,vtype=2)))
    update_menu.add_command(label="修改316L阀杆斤价", command=lambda:window.wait_window(inputDialog(window,'修改316L阀杆斤价','请输入新的316L阀杆斤价：',updrod_mcommand,vtype=3)))
    menu_bar.add_cascade(label="修改", menu=update_menu)
    inout_menu = Menu(menu_bar, tearoff=0)
    inout_menu.add_command(label="导入配件价格", command=input_mcommand)
    inout_menu.add_command(label="导出所有数据", command=output_mcommand)
    menu_bar.add_cascade(label="导入/导出", menu=inout_menu)
    about_menu = Menu(menu_bar, tearoff=0)
    about_menu.add_command(label="说明", command=lambda:messagebox.showinfo('说明','1.备注中的内容可以随意填写，一般‘；’前的内容体现在其他费用中。\n2.配件采用‘名称-规格-材质’命名，暂不可区分配件在规格中加‘+金额’表示。\n3.有成品有称斤的配件需要在名称后加‘成品’进行区分。\n4..数据保存在‘我的电脑-文档-阀门明细’中，其他主机可通过‘网络-Desktop-ihf5od1-阀门明细’查看。\n5.其他软件使用说明，见‘使用说明.txt’'))
    about_menu.add_command(label="关于我们", command=lambda:messagebox.showinfo('关于','天泰阀门内部使用，有问题请联系XXXXXXXXXXX。'))
    menu_bar.add_cascade(label="其他", menu=about_menu)                
    window.config(menu=menu_bar)

    frame = Frame(window)
    frame.grid(row=0, column=0, sticky='nsew')
    frame.grid_rowconfigure(0,minsize=56)
    frame.grid_rowconfigure(8,weight=1)
    frame.grid_columnconfigure(0, weight=1)
    lfont = Font(family='Microsoft YaHei', size=14)
    efont = Font(family='SimHei', size=12)

    frame1 = Frame(frame,height=56,width=763)
    frame1.grid(row=0, column=0, columnspan=2, rowspan=3, pady=7)
    l_name = Button(frame1, font=lfont, width=6, text="型号", command=select_valve)
    l_name.place(relx=0.08,rely=0.5,anchor='center',height=40)
    name = StringVar()
    E_name = Entry(frame1, font=efont, width=16, textvariable=name)
    E_name.place(relx=0.247,rely=0.5,anchor='center')
    E_name.bind("<KeyRelease>", autocomplete)
    E_name.bind("<Button-1>", autocomplete)
    E_name.bind("<Tab>", autocomplete_cancel)

    l_supply = Button(frame1, font=lfont, width=6, text="货源", command=show_compare)
    l_supply.place(relx=0.42,rely=0.5,anchor='center',height=40)
    supply = StringVar()
    E_supply = Entry(frame1, font=efont, width=14, textvariable=supply)
    E_supply.place(relx=0.587,rely=0.5,anchor='center')
    E_supply.bind("<Button-1>", lambda e: show_supply())
    B_supply_del = Label(frame1, text='×', font=Font(size=8), bg='white', cursor='hand2')
    B_supply_del.place(relx=0.653,rely=0.51,anchor='center',width=15,height=15)
    B_supply_del.bind("<ButtonRelease-1>", lambda e:del_supply())
    B_supply_list = Button(frame1, text='▼', font=Font(size=8), command=show_supply)
    B_supply_list.place(relx=0.67,rely=0.5,anchor='center',width=17,height=19)
    l_cost = Button(frame1, font=lfont, width=6, text="成本(元)", command=show_costdetails)
    l_cost.place(relx=0.767,rely=0.5,anchor='center',height=40)
    l_cost.bind("<Return>", lambda e:show_costdetails())
    l_cost.bind("<KP_Enter>", lambda e:show_costdetails())
    cost = StringVar()
    E_cost = Entry(frame1, font=efont, width=12, textvariable=cost)
    E_cost.place(relx=0.927,rely=0.5,anchor='center')
    E_cost.bind("<Tab>", lambda e:E_weight.focus_set())
    E_cost.bind("<Key>", lambda e:"break")#禁止手动修改成本

    frame2 = Frame(frame)
    frame2.grid(row=3, column=0, rowspan=5, padx=10)
    l_date = Label(frame2, font=lfont, width=12, text="更新日期")
    l_date.grid(row=0, column=0)
    date = StringVar()
    E_date = Entry(frame2, font=efont, width=12, textvariable=date)
    E_date.grid(row=0, column=1)
    l_weight = Label(frame2, font=lfont, width=12, text="成品重量(斤)")
    l_weight.grid(row=0, column=2)
    weight = StringVar()
    E_weight = Entry(frame2, font=efont, width=12, textvariable=weight)
    E_weight.grid(row=0, column=3)
    l_hnumber = Label(frame2, font=lfont, width=12, text="炉号")
    l_hnumber.grid(row=0, column=4)
    hnumber = StringVar()
    E_hnumber = Entry(frame2, font=efont, width=12, textvariable=hnumber)
    E_hnumber.grid(row=0, column=5)
    l_bweight = Label(frame2, font=lfont, width=12, text="毛坯重量(斤)")
    l_bweight.grid(row=1, column=0)
    bweight = StringVar()
    E_bweight = Entry(frame2, font=efont, width=12, textvariable=bweight)
    E_bweight.grid(row=1, column=1)
    l_bwcost = Label(frame2, font=lfont, width=12, text="毛坯单价")
    l_bwcost.grid(row=1, column=2)
    bwcost = StringVar()
    E_bwcost = Entry(frame2, font=efont, width=12, textvariable=bwcost)
    E_bwcost.grid(row=1, column=3)
    l_process = Label(frame2, font=lfont, width=12, text="加工费(元)")
    l_process.grid(row=1, column=4)
    process = StringVar()
    E_process = Entry(frame2, font=efont, width=12, textvariable=process)
    E_process.grid(row=1, column=5)
    l_rodd = Label(frame2, font=lfont, width=12, text="阀杆外圆")
    l_rodd.grid(row=2, column=0)
    rodd = StringVar()
    E_rodd = Entry(frame2, font=efont, width=12, textvariable=rodd)
    E_rodd.grid(row=2, column=1)
    l_rodl = Label(frame2, font=lfont, width=12, text="阀杆总长")
    l_rodl.grid(row=2, column=2)
    rodl = StringVar()
    E_rodl = Entry(frame2, font=efont, width=12, textvariable=rodl)
    E_rodl.grid(row=2, column=3)
    l_diff = Label(frame2, font=lfont, width=12, text="其他费用(元)")
    l_diff.grid(row=2, column=4)
    diff = StringVar()
    E_diff = Entry(frame2, font=efont, width=12, textvariable=diff)
    E_diff.grid(row=2, column=5)
    l_remark = Label(frame2, font=lfont, width=12, text="备注")
    l_remark.grid(row=3, column=0)
    E_remark = Text(frame2, font=efont, width=75, wrap=WORD, height=2)#font="TkTextFont"
    E_remark.grid(row=3, column=1, rowspan=2, columnspan=5, pady=6)
    E_remark.bind("<Tab>", lambda e:E_remark_tab())#强制Text无视制表符，而转移焦点
    E_remark.bind("<Shift-Tab>", lambda e:E_remark_stab())

    frame3 = Frame(frame)
    frame3.grid(row=8, column=0, rowspan=8, padx=30, pady=6, sticky='nsew')
    frame3.grid_rowconfigure(2,weight=1)
    frame3.grid_columnconfigure(4,minsize=30)
    frame3.grid_columnconfigure(0,weight=1)
    l9 = Label(frame3, font=lfont, text="配件明细", anchor='w')
    l9.grid(row=0, column=0, sticky='w')
    item_list = ttk.Treeview(frame3, columns=("名称","规格","材质","数量","单价","小计"), show="headings")
    style = ttk.Style()
    style.configure("Treeview.Heading", font=lfont)
    style.configure("Treeview", font=efont)
    item_list.heading("名称", text="名称", command=search_command)
    item_list.heading("规格", text="规格", command=search_command)
    item_list.heading("材质", text="材质", command=search_command)
    item_list.heading("数量", text="数量", command=lambda col='数量': sort_tree(item_list,col,True))
    item_list.heading("单价", text="单价", command=lambda col='单价': sort_tree(item_list,col,True))
    item_list.heading("小计", text="小计", command=lambda col='小计': sort_tree(item_list,col,True))
    item_list.column("名称", width=75)
    item_list.column("规格", width=90)
    item_list.column("材质", width=60)
    item_list.column("数量", width=60)
    item_list.column("单价", width=60)
    item_list.column("小计", width=60)
    item_list.grid(row=1, column=0, rowspan=8, columnspan=3, sticky='nsew')
    sb1 = Scrollbar(frame3)
    sb1.grid(row=1, column=3, rowspan=8, sticky='ns')
    item_list.configure(selectmode='browse', yscrollcommand=sb1.set)
    sb1.configure(command=item_list.yview, takefocus=False)
    item_list.bind("<<TreeviewSelect>>", get_selected_row)
    item_list.bind("<Left>",left2tv)
    item_list.bind("<Delete>",tv_delete)

    b2 = Button(frame3, font=lfont, text="查询", width=8, command=lambda:search_command(cursupply=''))
    b2.grid(row=0, rowspan=2, column=5, padx=10, pady=20)
    b2.bind("<Return>", lambda e:search_command(cursupply=''))
    b2.bind("<KP_Enter>", lambda e:search_command(cursupply=''))
    b3 = Button(frame3, font=lfont, text="修改", width=8, command=update_command)
    b3.grid(row=0, rowspan=2, column=6, padx=10, pady=20)
    b3.bind("<Return>", lambda e:update_command())
    b3.bind("<KP_Enter>", lambda e:update_command())

    F_iname = Frame(frame3)
    F_iname.grid(row=4, column=5, rowspan=2, columnspan=2, pady=20)
    l_iname = Label(F_iname, font=lfont, text="配件名称-规格-材质", anchor='w', width=20)
    l_iname.grid(row=0, column=1, columnspan=4, pady=10, sticky='w')
    F_iname.grid_columnconfigure(0,minsize=16)
    F_iname.grid_columnconfigure(1,minsize=39)
    iname_type=StringVar()
    E_iname_type = SuperEntry(F_iname, placeholder='名称', font=efont, textvariable=iname_type, width=8)
    E_iname_type.grid(row=1, column=2)
    iname_spe=StringVar()
    E_iname_spe = SuperEntry(F_iname, placeholder='规格', font=efont, textvariable=iname_spe, width=12)
    E_iname_spe.grid(row=1, column=3)
    iname_tex=StringVar()
    E_iname_tex = SuperEntry(F_iname, placeholder='材质', font=efont, textvariable=iname_tex, width=8)
    E_iname_tex.grid(row=1, column=4)
    E_iname_type.bind("<KeyRelease>", autocomplete)
    E_iname_type.bind("<Button-1>", autocomplete)
    E_iname_type.bind("<Left>", left2tv)
    E_iname_type.bind("<Right>", lambda e:iname_direct(e,E_iname_spe,0))
    E_iname_spe.bind("<KeyRelease>", autocomplete)
    E_iname_spe.bind("<Button-1>", autocomplete)
    E_iname_spe.bind("<Right>", lambda e:iname_direct(e,E_iname_tex,0))
    E_iname_spe.bind("<Left>", lambda e:iname_direct(e,E_iname_type,1))
    E_iname_tex.bind("<KeyRelease>", autocomplete)
    E_iname_tex.bind("<Button-1>", autocomplete)
    E_iname_tex.bind("<Left>", lambda e:iname_direct(e,E_iname_spe,1))
    E_iname_tex.bind("<Tab>", lambda e:(autocomplete_select(autocomplete_listbox, E_iname_tex, autocomplete_listbox.get(0)) if 'autocomplete_listbox' in globals() and autocomplete_listbox.winfo_exists() else None, autocomplete_cancel(e)))
    E_iname_tex.bind("<Right>",lambda e:(E_icost.delete(0,END),iname_direct(e,E_icost,0),autocomplete_toplevel.destroy() if 'autocomplete_listbox' in globals() else None))

    l_icost = Label(frame3, font=lfont, text="配件价格")
    l_icost.grid(row=6, column=5, padx=7, pady=20)
    icost = StringVar()
    E_icost = Entry(frame3, font=efont, textvariable=icost)
    E_icost.grid(row=6, column=6, padx=7, pady=20)
    E_icost.bind("<Left>", left2tv)
    item_list.bind("<Right>", tv_right)

    l_icount = Label(frame3, font=lfont, text="配件数量")
    l_icount.grid(row=7, column=5, padx=7, pady=20)
    icount = StringVar()
    E_icount = Entry(frame3, font=efont, textvariable=icount)
    E_icount.grid(row=7, column=6, padx=7, pady=20)
    E_icount.bind("<Left>", left2tv)

    b4 = Button(frame3, font=lfont, text="添加配件", width=8, command=additem_command)
    b4.grid(row=8, column=5, padx=10, pady=20)
    b4.bind("<Return>", lambda e:additem_command())
    b4.bind("<KP_Enter>", lambda e:additem_command())
    b5 = Button(frame3, font=lfont, text="删除配件", width=8, command=delitem_command)
    b5.grid(row=8, column=6, padx=10, pady=20)
    b5.bind("<Return>", lambda e:delitem_command())
    b5.bind("<KP_Enter>", lambda e:delitem_command())

    window.update_idletasks()##更新窗口信息以使窗口居中显示
    window_x = (window.winfo_screenwidth() // 2) - (window.winfo_width() // 2)
    window_y = int(window.winfo_screenheight() // 2.1) - (window.winfo_height() // 2)
    window.geometry(f"+{window_x}+{window_y}")
    window.focus_force()
    E_name.focus_set()
    window.mainloop()



    ## nextlevetl:
    # 查看阀门，展示多个同型号不同来源（后缀）不同货源的阀门及其各部分成本价
